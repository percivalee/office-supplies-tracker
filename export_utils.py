from datetime import datetime
from io import BytesIO
from typing import Any, Iterable, Mapping, Optional
from urllib.parse import quote

EXPORT_HEADERS = (
    "流水号", "申领日期", "申领部门", "经办人", "物品名称",
    "数量", "单价", "状态", "到货日期", "分发对象", "分发日期", "签收备注",
)
EXPORT_COLUMN_WIDTHS = (18, 12, 24, 12, 28, 10, 10, 12, 12, 14, 12, 28)
EXPORT_FALLBACK_FILENAME = "office_supplies_export.xlsx"
EXPORT_DISPLAY_NAME_PREFIX = "办公用品台账"


class ExportDependencyError(RuntimeError):
    """Export dependency is unavailable."""


def _build_item_row(item: Mapping[str, Any]) -> list[Any]:
    return [
        item.get("serial_number", ""),
        item.get("request_date", ""),
        item.get("department", ""),
        item.get("handler", ""),
        item.get("item_name", ""),
        item.get("quantity", ""),
        "" if item.get("unit_price") is None else item.get("unit_price"),
        item.get("status", ""),
        item.get("arrival_date", ""),
        item.get("recipient", ""),
        item.get("distribution_date", ""),
        item.get("signoff_note", ""),
    ]


def build_items_excel_stream(items: Iterable[Mapping[str, Any]]) -> BytesIO:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as exc:
        raise ExportDependencyError("缺少 openpyxl 依赖，请先安装 requirements.txt") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "采购记录"
    sheet.append(list(EXPORT_HEADERS))

    for item in items:
        sheet.append(_build_item_row(item))

    for idx, width in enumerate(EXPORT_COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def build_export_content_disposition(now: Optional[datetime] = None) -> str:
    timestamp = (now or datetime.now()).strftime("%Y%m%d_%H%M%S")
    filename = f"{EXPORT_DISPLAY_NAME_PREFIX}_{timestamp}.xlsx"
    encoded_filename = quote(filename)
    return (
        f"attachment; filename=\"{EXPORT_FALLBACK_FILENAME}\"; "
        f"filename*=UTF-8''{encoded_filename}"
    )
