from contextlib import asynccontextmanager
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import HTMLResponse, FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field
from typing import Optional
import aiosqlite
import shutil
import re
import sys
import asyncio
import zipfile
from pathlib import Path
from io import BytesIO
from datetime import datetime
from uuid import uuid4
from starlette.concurrency import run_in_threadpool

from database import (
    init_db, get_items, count_items, get_stats_summary, get_item, create_item, update_item, delete_item,
    batch_create_items, get_serial_numbers, get_departments, get_handlers,
    ItemStatus, PaymentStatus, DB_PATH
)
from parser import parse_document


@asynccontextmanager
async def lifespan(app: FastAPI):
    """启动时初始化数据库"""
    await init_db()
    yield


app = FastAPI(title="办公用品采购追踪系统", lifespan=lifespan)

ALLOWED_UPLOAD_EXTENSIONS = {".pdf", ".png", ".jpg", ".jpeg", ".jfif"}
BACKUP_RESTORE_LOCK = asyncio.Lock()

def _resolve_runtime_dir() -> Path:
    """运行目录（源码模式为项目目录，打包模式为 exe 所在目录）。"""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def _resolve_static_dir() -> Path:
    """静态资源目录（兼容源码与 PyInstaller 打包）。"""
    candidates: list[Path] = []
    meipass = getattr(sys, "_MEIPASS", None)
    if meipass:
        candidates.append(Path(meipass) / "static")
    candidates.append(Path(__file__).resolve().parent / "static")
    candidates.append(Path.cwd() / "static")

    for path in candidates:
        if path.exists():
            return path
    return candidates[0]


RUNTIME_DIR = _resolve_runtime_dir()
STATIC_DIR = _resolve_static_dir()

# 创建上传目录（落在运行目录，避免打包后写入只读目录）
UPLOAD_DIR = RUNTIME_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

# 挂载静态文件
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")


def _item_key(item: dict) -> tuple[str, str, str]:
    """构造判重键：流水号 + 物品名称 + 经办人。"""
    return (
        str(item.get("serial_number") or "").strip(),
        str(item.get("item_name") or "").strip(),
        str(item.get("handler") or "").strip(),
    )


def _safe_quantity(value) -> float:
    """数量兜底，避免异常值导致合并失败。"""
    try:
        qty = float(value)
        return qty if qty > 0 else 1.0
    except (TypeError, ValueError):
        return 1.0


def _normalize_month(month: Optional[str]) -> Optional[str]:
    """校验月份参数，格式为 YYYY-MM。"""
    if month is None:
        return None
    month = month.strip()
    if not month:
        return None
    if not re.fullmatch(r"\d{4}-(0[1-9]|1[0-2])", month):
        raise HTTPException(status_code=400, detail="month 参数格式应为 YYYY-MM")
    return month


def _normalize_text_filter(value: Optional[str]) -> Optional[str]:
    """将空白筛选值归一化为 None。"""
    if value is None:
        return None
    value = value.strip()
    return value or None


def _safe_unlink(path: Path) -> None:
    """安全删除临时文件。"""
    try:
        path.unlink(missing_ok=True)
    except OSError:
        pass


def _build_upload_path(filename: str) -> Path:
    """构造唯一上传路径，并校验扩展名。"""
    safe_filename = Path(filename).name if filename else ""
    if not safe_filename:
        raise HTTPException(status_code=400, detail="无效的文件名")

    extension = Path(safe_filename).suffix.lower()
    if extension not in ALLOWED_UPLOAD_EXTENSIONS:
        raise HTTPException(status_code=400, detail="仅支持 PDF / PNG / JPG / JPEG / JFIF 文件")

    unique_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}_{uuid4().hex}{extension}"
    return UPLOAD_DIR / unique_name


def _resolve_db_path() -> Path:
    """解析数据库路径（兼容相对路径配置）。"""
    db_path = Path(DB_PATH)
    if db_path.is_absolute():
        return db_path
    return RUNTIME_DIR / db_path


def _is_safe_zip_entry(name: str) -> bool:
    """校验压缩包内路径，阻止目录穿越。"""
    path = Path(name)
    if path.is_absolute():
        return False
    return ".." not in path.parts


def _build_backup_archive() -> tuple[BytesIO, str]:
    """打包数据库与上传目录为 zip。"""
    buffer = BytesIO()
    filename = f"office_supplies_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
    db_path = _resolve_db_path()

    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        if db_path.exists():
            archive.write(db_path, arcname="office_supplies.db")
        if UPLOAD_DIR.exists():
            for file_path in UPLOAD_DIR.rglob("*"):
                if file_path.is_file():
                    arcname = Path("uploads") / file_path.relative_to(UPLOAD_DIR)
                    archive.write(file_path, arcname=arcname.as_posix())

    buffer.seek(0)
    return buffer, filename


def _restore_from_archive(archive_path: Path) -> dict:
    """从备份 zip 恢复数据库与上传目录。"""
    extract_dir = RUNTIME_DIR / f".restore_tmp_{uuid4().hex}"
    snapshot_db = RUNTIME_DIR / f".restore_db_snapshot_{uuid4().hex}.bak"
    snapshot_uploads = RUNTIME_DIR / f".restore_uploads_snapshot_{uuid4().hex}"
    db_path = _resolve_db_path()

    try:
        extract_dir.mkdir(parents=True, exist_ok=False)

        with zipfile.ZipFile(archive_path, "r") as archive:
            members = [name for name in archive.namelist() if name and not name.endswith("/")]
            if not members:
                raise ValueError("备份包为空")
            for name in members:
                if not _is_safe_zip_entry(name):
                    raise ValueError("备份包包含非法路径")
            archive.extractall(extract_dir)

        restored_db = extract_dir / "office_supplies.db"
        restored_uploads = extract_dir / "uploads"
        if not restored_db.exists():
            raise ValueError("备份包缺少 office_supplies.db")

        if db_path.exists():
            snapshot_db.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(db_path, snapshot_db)
        if UPLOAD_DIR.exists():
            shutil.copytree(UPLOAD_DIR, snapshot_uploads)

        try:
            db_path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(restored_db, db_path)

            shutil.rmtree(UPLOAD_DIR, ignore_errors=True)
            UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

            restored_upload_files = 0
            if restored_uploads.exists():
                for src_file in restored_uploads.rglob("*"):
                    if not src_file.is_file():
                        continue
                    relative = src_file.relative_to(restored_uploads)
                    dest_file = UPLOAD_DIR / relative
                    dest_file.parent.mkdir(parents=True, exist_ok=True)
                    shutil.copy2(src_file, dest_file)
                    restored_upload_files += 1

            if not any(UPLOAD_DIR.iterdir()):
                (UPLOAD_DIR / ".gitkeep").touch(exist_ok=True)

            return {
                "restored_db": True,
                "restored_upload_files": restored_upload_files,
            }
        except Exception:
            if snapshot_db.exists():
                shutil.copy2(snapshot_db, db_path)
            if snapshot_uploads.exists():
                shutil.rmtree(UPLOAD_DIR, ignore_errors=True)
                shutil.copytree(snapshot_uploads, UPLOAD_DIR)
            raise
    finally:
        shutil.rmtree(extract_dir, ignore_errors=True)
        _safe_unlink(snapshot_db)
        shutil.rmtree(snapshot_uploads, ignore_errors=True)


def _normalize_import_payload(payload: dict) -> dict:
    """标准化导入数据，合并同键明细并清理空值。"""
    serial_number = str(payload.get("serial_number") or "").strip()
    department = str(payload.get("department") or "").strip()
    handler = str(payload.get("handler") or "").strip()
    request_date = str(payload.get("request_date") or "").strip()

    merged_items: dict[tuple[str, str, str], dict] = {}
    for raw in payload.get("items", []):
        item_name = str((raw or {}).get("item_name") or "").strip()
        if not item_name:
            continue
        key = (serial_number, item_name, handler)
        quantity = _safe_quantity((raw or {}).get("quantity"))
        purchase_link_raw = str((raw or {}).get("purchase_link") or "").strip()
        purchase_link = purchase_link_raw or None

        if key in merged_items:
            merged_items[key]["quantity"] += quantity
            if not merged_items[key].get("purchase_link") and purchase_link:
                merged_items[key]["purchase_link"] = purchase_link
            continue

        merged_items[key] = {
            "serial_number": serial_number,
            "department": department,
            "handler": handler,
            "request_date": request_date,
            "item_name": item_name,
            "quantity": quantity,
            "purchase_link": purchase_link,
        }

    return {
        "serial_number": serial_number,
        "department": department,
        "handler": handler,
        "request_date": request_date,
        "items": list(merged_items.values()),
    }


def _build_preview_data(normalized_payload: dict, items: list[dict]) -> dict:
    """构造可回显到前端的预览数据。"""
    return {
        "serial_number": normalized_payload.get("serial_number", ""),
        "department": normalized_payload.get("department", ""),
        "handler": normalized_payload.get("handler", ""),
        "request_date": normalized_payload.get("request_date", ""),
        "items": [
            {
                "item_name": item.get("item_name", ""),
                "quantity": _safe_quantity(item.get("quantity")),
                "purchase_link": item.get("purchase_link"),
            }
            for item in items
        ],
    }


def _collect_duplicates(items: list[dict], existing_by_key: dict) -> list[dict]:
    """收集与数据库已存在记录冲突的明细。"""
    duplicates: list[dict] = []
    for item in items:
        key = _item_key(item)
        existing = existing_by_key.get(key)
        if not existing:
            continue
        duplicates.append({
            "serial_number": item["serial_number"],
            "item_name": item["item_name"],
            "handler": item["handler"],
            "existing_quantity": existing["quantity"],
            "new_quantity": item["quantity"],
            "department": item["department"],
            "existing_id": existing["id"]
        })
    return duplicates


async def _confirm_import(normalized_payload: dict, duplicate_action: Optional[str] = None) -> dict:
    """根据导入数据执行确认导入流程。"""
    items_to_create = normalized_payload.get("items", [])
    if not items_to_create:
        raise HTTPException(status_code=400, detail="未识别到可导入的物品明细")

    if duplicate_action is not None and duplicate_action not in {"skip", "add", "merge"}:
        raise HTTPException(status_code=400, detail="不支持的操作类型")

    all_items = await get_items()
    existing_by_key = {_item_key(item): item for item in all_items}
    duplicates = _collect_duplicates(items_to_create, existing_by_key)
    preview_data = _build_preview_data(normalized_payload, items_to_create)

    if duplicates and duplicate_action is None:
        return {
            "message": f"检测到 {len(duplicates)} 个重复物品，请选择处理方式",
            "has_duplicates": True,
            "duplicates": duplicates,
            "parsed_data": preview_data,
        }

    action = duplicate_action or "add"
    created_ids: list[int] = []
    updated_count = 0
    to_insert: list[dict] = []

    if action == "skip":
        to_insert = [item for item in items_to_create if _item_key(item) not in existing_by_key]
        if to_insert:
            created_ids = await batch_create_items(to_insert)
    elif action == "add":
        to_insert = list(items_to_create)
        created_ids = await batch_create_items(to_insert)
    elif action == "merge":
        quantity_updates: dict[int, float] = {}
        for item in items_to_create:
            item_key = _item_key(item)
            qty = _safe_quantity(item.get("quantity"))
            existing = existing_by_key.get(item_key)
            if existing:
                existing_id = existing["id"]
                base_qty = quantity_updates.get(existing_id, _safe_quantity(existing["quantity"]))
                quantity_updates[existing_id] = base_qty + qty
            else:
                new_item = dict(item)
                new_item["quantity"] = qty
                to_insert.append(new_item)

        if to_insert:
            created_ids = await batch_create_items(to_insert)
        for item_id, quantity in quantity_updates.items():
            await update_item(item_id, {"quantity": quantity})
        updated_count = len(quantity_updates)

    return {
        "message": f"导入完成：新增 {len(created_ids)} 条，更新 {updated_count} 条",
        "has_duplicates": False,
        "parsed_data": preview_data,
        "created_count": len(created_ids),
        "created_ids": created_ids,
        "updated_count": updated_count,
    }


# Pydantic 模型
class ItemCreate(BaseModel):
    serial_number: str
    department: str
    handler: str
    request_date: str
    item_name: str
    quantity: float = Field(gt=0)
    purchase_link: Optional[str] = None
    unit_price: Optional[float] = Field(default=None, ge=0)
    status: str = "待采购"
    invoice_issued: bool = False
    payment_status: str = "未付款"


class ItemUpdate(BaseModel):
    serial_number: Optional[str] = None
    department: Optional[str] = None
    handler: Optional[str] = None
    request_date: Optional[str] = None
    item_name: Optional[str] = None
    quantity: Optional[float] = Field(default=None, gt=0)
    purchase_link: Optional[str] = None
    unit_price: Optional[float] = Field(default=None, ge=0)
    status: Optional[str] = None
    invoice_issued: Optional[bool] = None
    payment_status: Optional[str] = None


class ImportItem(BaseModel):
    item_name: str = ""
    quantity: Optional[float] = None
    purchase_link: Optional[str] = None


class ImportConfirmRequest(BaseModel):
    serial_number: str = ""
    department: str = ""
    handler: str = ""
    request_date: str = ""
    items: list[ImportItem] = Field(default_factory=list)
    duplicate_action: Optional[str] = None


@app.get("/", response_class=HTMLResponse)
async def root():
    """返回主页"""
    html_path = STATIC_DIR / "index.html"
    return FileResponse(html_path)


@app.get("/api/items")
async def list_items(
    status: Optional[str] = None,
    department: Optional[str] = None,
    month: Optional[str] = None,
    keyword: Optional[str] = None,
    page: int = 1,
    page_size: int = 20
):
    """获取所有物品列表"""
    if page < 1:
        raise HTTPException(status_code=400, detail="page 必须 >= 1")
    if page_size < 1 or page_size > 200:
        raise HTTPException(status_code=400, detail="page_size 必须在 1-200 之间")

    status = _normalize_text_filter(status)
    department = _normalize_text_filter(department)
    month = _normalize_month(month)
    keyword = _normalize_text_filter(keyword)
    items = await get_items(
        status=status, department=department, month=month, keyword=keyword,
        page=page, page_size=page_size
    )
    total = await count_items(status=status, department=department, month=month, keyword=keyword)
    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size
    }


@app.get("/api/export")
async def export_items(
    status: Optional[str] = None,
    department: Optional[str] = None,
    month: Optional[str] = None,
    keyword: Optional[str] = None
):
    """导出筛选后的记录为 Excel"""
    status = _normalize_text_filter(status)
    department = _normalize_text_filter(department)
    month = _normalize_month(month)
    keyword = _normalize_text_filter(keyword)
    items = await get_items(status=status, department=department, month=month, keyword=keyword)

    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError:
        raise HTTPException(status_code=500, detail="缺少 openpyxl 依赖，请先安装 requirements.txt")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "采购记录"

    headers = ["流水号", "申领日期", "申领部门", "经办人", "物品名称", "数量", "单价", "状态"]
    sheet.append(headers)

    for item in items:
        sheet.append([
            item.get("serial_number", ""),
            item.get("request_date", ""),
            item.get("department", ""),
            item.get("handler", ""),
            item.get("item_name", ""),
            item.get("quantity", ""),
            "" if item.get("unit_price") is None else item.get("unit_price"),
            item.get("status", ""),
        ])

    column_widths = [18, 12, 24, 12, 28, 10, 10, 12]
    for idx, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"office_supplies_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@app.get("/api/backup")
async def backup_data():
    """下载当前数据库与上传文件备份。"""
    async with BACKUP_RESTORE_LOCK:
        try:
            archive_buffer, filename = await run_in_threadpool(_build_backup_archive)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"备份失败: {str(e)}")
    return StreamingResponse(
        archive_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@app.post("/api/restore")
async def restore_data(file: UploadFile = File(...)):
    """从备份包恢复数据库与上传文件。"""
    filename = (file.filename or "").strip()
    if not filename:
        raise HTTPException(status_code=400, detail="无效的备份文件名")
    if Path(filename).suffix.lower() != ".zip":
        raise HTTPException(status_code=400, detail="仅支持 .zip 备份文件")

    archive_path = UPLOAD_DIR / f"restore_{uuid4().hex}.zip"
    with open(archive_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    await file.close()

    async with BACKUP_RESTORE_LOCK:
        try:
            result = await run_in_threadpool(_restore_from_archive, archive_path)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"恢复失败: {str(e)}")
        finally:
            _safe_unlink(archive_path)

    return {
        "message": f"恢复完成，已恢复数据库与 {result['restored_upload_files']} 个上传文件",
        "restored_upload_files": result["restored_upload_files"],
    }


@app.get("/api/items/{item_id}")
async def read_item(item_id: int):
    """获取单个物品详情"""
    item = await get_item(item_id)
    if not item:
        raise HTTPException(status_code=404, detail="物品不存在")
    return item


@app.post("/api/items")
async def create_new_item(item: ItemCreate):
    """手动创建新物品记录"""
    try:
        item_id = await create_item(item.model_dump())
    except aiosqlite.IntegrityError:
        raise HTTPException(status_code=409, detail="记录已存在（流水号+物品名称+经办人）")
    return {"id": item_id, "message": "创建成功"}


@app.put("/api/items/{item_id}")
async def update_item_endpoint(item_id: int, updates: ItemUpdate):
    """更新物品记录"""
    update_data = updates.model_dump(exclude_unset=True)
    if not update_data:
        raise HTTPException(status_code=400, detail="未提供可更新字段")
    if "quantity" in update_data and update_data["quantity"] is None:
        raise HTTPException(status_code=400, detail="quantity 不能为空")
    try:
        success = await update_item(item_id, update_data)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except aiosqlite.IntegrityError as e:
        if "UNIQUE constraint failed" in str(e):
            raise HTTPException(status_code=409, detail="记录已存在（流水号+物品名称+经办人）")
        raise HTTPException(status_code=400, detail="更新失败：字段值不合法")
    if not success:
        raise HTTPException(status_code=404, detail="物品不存在")
    return {"message": "更新成功"}


@app.delete("/api/items/{item_id}")
async def delete_item_endpoint(item_id: int):
    """删除物品记录"""
    success = await delete_item(item_id)
    if not success:
        raise HTTPException(status_code=404, detail="物品不存在")
    return {"message": "删除成功"}


@app.post("/api/upload")
async def upload_and_parse(file: UploadFile = File(...)):
    """上传文件并解析"""
    file_path = _build_upload_path(file.filename or "")

    # 保存文件（使用唯一临时名，避免同名覆盖）
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    await file.close()

    try:
        # 解析文档（避免阻塞事件循环）
        result = await run_in_threadpool(parse_document, str(file_path))
        normalized_payload = _normalize_import_payload({
            "serial_number": result.get("serial_number", ""),
            "department": result.get("department", ""),
            "handler": result.get("handler", ""),
            "request_date": result.get("request_date", ""),
            "items": result.get("items", []),
        })
        preview_data = _build_preview_data(normalized_payload, normalized_payload["items"])

        return {
            "message": f"解析完成，共 {len(preview_data['items'])} 条，请确认后导入",
            "parsed_data": preview_data,
            "has_duplicates": False,
            "requires_confirmation": True,
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"解析失败: {str(e)}")
    finally:
        _safe_unlink(file_path)


@app.post("/api/import/confirm")
async def confirm_import(request: ImportConfirmRequest):
    """确认导入（支持人工校正后提交）。"""
    payload = request.model_dump()
    duplicate_action = payload.pop("duplicate_action", None)
    normalized_payload = _normalize_import_payload(payload)
    return await _confirm_import(normalized_payload, duplicate_action)


@app.get("/api/autocomplete")
async def autocomplete():
    """获取自动补全数据"""
    return {
        "serial_numbers": await get_serial_numbers(),
        "departments": await get_departments(),
        "handlers": await get_handlers(),
        "statuses": [s.value for s in ItemStatus],
        "payment_statuses": [s.value for s in PaymentStatus]
    }


@app.get("/api/stats")
async def get_stats():
    """获取统计信息"""
    return await get_stats_summary()


class DuplicateHandleRequest(BaseModel):
    """处理重复物品的请求"""
    action: str  # 'skip', 'add', 'merge'
    duplicates: list[dict]
    items_data: list[dict]  # 所有待插入的数据


@app.post("/api/upload/handle-duplicates")
async def handle_duplicates(request: DuplicateHandleRequest):
    """兼容旧前端：处理重复物品"""
    try:
        first = request.items_data[0] if request.items_data else {}
        normalized_payload = _normalize_import_payload({
            "serial_number": first.get("serial_number", ""),
            "department": first.get("department", ""),
            "handler": first.get("handler", ""),
            "request_date": first.get("request_date", ""),
            "items": request.items_data,
        })
        return await _confirm_import(normalized_payload, request.action)

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"处理失败: {str(e)}")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
