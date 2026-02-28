from datetime import datetime
from io import BytesIO
from typing import Optional
from urllib.parse import quote

import aiosqlite
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse

from api_utils import normalize_history_action, normalize_month, normalize_text_filter
from database import (
    ItemStatus,
    PaymentStatus,
    batch_update_items,
    count_item_history,
    count_items,
    create_item,
    delete_item,
    get_amount_report,
    get_departments,
    get_handlers,
    get_item,
    get_item_history,
    get_items,
    get_serial_numbers,
    get_stats_summary,
    update_item,
)
from schemas import BatchUpdateRequest, ItemCreate, ItemUpdate

router = APIRouter(prefix="/api")
MIN_PAGE = 1
MAX_PAGE_SIZE = 200
DEFAULT_PAGE_SIZE = 20


def _validate_pagination(page: int, page_size: int) -> None:
    if page < MIN_PAGE:
        raise HTTPException(status_code=400, detail="page 必须 >= 1")
    if page_size < 1 or page_size > MAX_PAGE_SIZE:
        raise HTTPException(status_code=400, detail="page_size 必须在 1-200 之间")


def _normalize_item_filters(
    status: Optional[str],
    department: Optional[str],
    month: Optional[str],
    keyword: Optional[str],
) -> tuple[Optional[str], Optional[str], Optional[str], Optional[str]]:
    return (
        normalize_text_filter(status),
        normalize_text_filter(department),
        normalize_month(month),
        normalize_text_filter(keyword),
    )


def _normalize_history_filters(
    action: Optional[str],
    keyword: Optional[str],
    month: Optional[str],
) -> tuple[Optional[str], Optional[str], Optional[str]]:
    return (
        normalize_history_action(action),
        normalize_text_filter(keyword),
        normalize_month(month),
    )


@router.get("/items")
async def list_items(
    status: Optional[str] = None,
    department: Optional[str] = None,
    month: Optional[str] = None,
    keyword: Optional[str] = None,
    page: int = 1,
    page_size: int = DEFAULT_PAGE_SIZE
):
    """获取所有物品列表。"""
    _validate_pagination(page, page_size)
    status, department, month, keyword = _normalize_item_filters(
        status, department, month, keyword
    )
    items = await get_items(
        status=status, department=department, month=month, keyword=keyword,
        page=page, page_size=page_size
    )
    total = await count_items(status=status, department=department, month=month, keyword=keyword)
    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@router.post("/items/batch-update")
async def batch_update_items_endpoint(request: BatchUpdateRequest):
    """批量更新记录。"""
    if not request.updates:
        raise HTTPException(status_code=400, detail="updates 不能为空")
    try:
        result = await batch_update_items(request.ids, request.updates)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except aiosqlite.IntegrityError as e:
        if "UNIQUE constraint failed" in str(e):
            raise HTTPException(status_code=409, detail="批量更新触发唯一约束冲突（流水号+物品名称+经办人）")
        raise HTTPException(status_code=400, detail="批量更新失败：字段值不合法")

    updated_count = result.get("updated_count", 0)
    unchanged_count = result.get("unchanged_count", 0)
    missing_ids = result.get("missing_ids", [])
    message = f"批量更新完成：更新 {updated_count} 条"
    if unchanged_count:
        message += f"，未变化 {unchanged_count} 条"
    if missing_ids:
        message += f"，未找到 {len(missing_ids)} 条"
    return {
        "message": message,
        **result,
    }


@router.get("/export")
async def export_items(
    status: Optional[str] = None,
    department: Optional[str] = None,
    month: Optional[str] = None,
    keyword: Optional[str] = None
):
    """导出筛选后的记录为 Excel。"""
    status, department, month, keyword = _normalize_item_filters(
        status, department, month, keyword
    )
    items = await get_items(status=status, department=department, month=month, keyword=keyword)

    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError:
        raise HTTPException(status_code=500, detail="缺少 openpyxl 依赖，请先安装 requirements.txt")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "采购记录"

    headers = ["流水号", "申领日期", "申领部门", "经办人", "物品名称", "数量", "单价", "状态"]
    sheet.append(headers)

    for item in items:
        sheet.append([
            item.get("serial_number", ""),
            item.get("request_date", ""),
            item.get("department", ""),
            item.get("handler", ""),
            item.get("item_name", ""),
            item.get("quantity", ""),
            "" if item.get("unit_price") is None else item.get("unit_price"),
            item.get("status", ""),
        ])

    column_widths = [18, 12, 24, 12, 28, 10, 10, 12]
    for idx, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"办公用品台账_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    encoded_filename = quote(filename)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


@router.get("/items/{item_id}")
async def read_item(item_id: int):
    """获取单个物品详情。"""
    item = await get_item(item_id)
    if not item:
        raise HTTPException(status_code=404, detail="物品不存在")
    return item


@router.post("/items")
async def create_new_item(item: ItemCreate):
    """手动创建新物品记录。"""
    try:
        item_id = await create_item(item.model_dump())
    except aiosqlite.IntegrityError:
        raise HTTPException(status_code=409, detail="记录已存在（流水号+物品名称+经办人）")
    return {"id": item_id, "message": "创建成功"}


@router.put("/items/{item_id}")
async def update_item_endpoint(item_id: int, updates: ItemUpdate):
    """更新物品记录。"""
    update_data = updates.model_dump(exclude_unset=True)
    if not update_data:
        raise HTTPException(status_code=400, detail="未提供可更新字段")
    if "quantity" in update_data and update_data["quantity"] is None:
        raise HTTPException(status_code=400, detail="quantity 不能为空")
    try:
        success = await update_item(item_id, update_data)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except aiosqlite.IntegrityError as e:
        if "UNIQUE constraint failed" in str(e):
            raise HTTPException(status_code=409, detail="记录已存在（流水号+物品名称+经办人）")
        raise HTTPException(status_code=400, detail="更新失败：字段值不合法")
    if not success:
        raise HTTPException(status_code=404, detail="物品不存在")
    return {"message": "更新成功"}


@router.delete("/items/{item_id}")
async def delete_item_endpoint(item_id: int):
    """删除物品记录。"""
    success = await delete_item(item_id)
    if not success:
        raise HTTPException(status_code=404, detail="物品不存在")
    return {"message": "删除成功"}


@router.get("/autocomplete")
async def autocomplete():
    """获取自动补全数据。"""
    return {
        "serial_numbers": await get_serial_numbers(),
        "departments": await get_departments(),
        "handlers": await get_handlers(),
        "statuses": [s.value for s in ItemStatus],
        "payment_statuses": [s.value for s in PaymentStatus],
    }


@router.get("/stats")
async def get_stats():
    """获取统计信息。"""
    return await get_stats_summary()


@router.get("/reports/amount")
async def amount_report(
    status: Optional[str] = None,
    department: Optional[str] = None,
    month: Optional[str] = None,
    keyword: Optional[str] = None
):
    """金额统计报表（支持与列表一致的筛选）。"""
    status, department, month, keyword = _normalize_item_filters(
        status, department, month, keyword
    )
    return await get_amount_report(
        status=status, department=department, month=month, keyword=keyword
    )


@router.get("/history")
async def history_list(
    action: Optional[str] = None,
    keyword: Optional[str] = None,
    month: Optional[str] = None,
    page: int = 1,
    page_size: int = DEFAULT_PAGE_SIZE
):
    """变更历史列表。"""
    _validate_pagination(page, page_size)
    action, keyword, month = _normalize_history_filters(action, keyword, month)
    items = await get_item_history(
        action=action, keyword=keyword, month=month,
        page=page, page_size=page_size
    )
    total = await count_item_history(action=action, keyword=keyword, month=month)
    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
    }
